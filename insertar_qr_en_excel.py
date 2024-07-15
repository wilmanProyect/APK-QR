import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

def insertar_imagenes_en_excel(excel_path, hoja, fila_inicio, fila_fin, columna_qr, carpeta_qrs, columna_codigo):
    wb = load_workbook(excel_path, data_only=True)  # Abrir con data_only=True para obtener los valores calculados
    ws = wb[hoja]
    
    for i in range(fila_inicio, fila_fin + 1):
        code = ws.cell(row=i, column=columna_codigo).value
        if code is None:
            print(f"No se encontró valor en la celda de código en la fila {i}")
            continue
        img_path = os.path.join(carpeta_qrs, f"{code}.png")
        if os.path.exists(img_path):
            img = XLImage(img_path)
            cell = ws.cell(row=i, column=columna_qr)
            ws.add_image(img, cell.coordinate)
        else:
            print(f"Imagen no encontrada: {img_path}")
    
    wb.save(excel_path)

if __name__ == "__main__":
    excel_path = r"D:\MyWork\APK QR\planilla de almendras para QR.xlsx"  # Reemplaza con la ruta a tu archivo Excel
    hoja = "Carmencita (3)"  # Reemplaza con el nombre de la hoja
    fila_inicio = 10  # Fila de inicio
    fila_fin = 179  # Fila de fin
    columna_qr = 9  # Columna donde se insertarán las imágenes (por ejemplo, 9 para la columna I)
    carpeta_qrs = r"D:\MyWork\APK QR\qr_generados"  # Carpeta donde se guardaron los QR
    columna_codigo = 8  # Columna que contiene los códigos QR (por ejemplo, 8 para la columna H)

    insertar_imagenes_en_excel(excel_path, hoja, fila_inicio, fila_fin, columna_qr, carpeta_qrs, columna_codigo)
#